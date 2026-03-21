# Copyright (C) 2023 - 2026 ANSYS, Inc. and/or its affiliates.
# SPDX-License-Identifier: MIT
#
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
PyEDB Local Design Diff & Regression Tool
=========================================
Compares two EDB designs (baseline vs. modified) and produces a rich
diff / regression report covering:

  ✔ Net connectivity  – added / removed / pin-count / primitives
  ✔ Components        – added / removed / moved / part-change / value-change
  ✔ Layer stackup     – thickness / material / Er / loss-tangent / order
  ✔ Padstack defs     – drill size / pad geometry per layer
  ✔ Trace geometry    – width / length / layer reassignment per net
  ✔ Via instances     – count / location deltas per net
  ✔ S-param regression – IL / RL / impedance vs. configurable thresholds
  ✔ JSON + Excel output with colour-coded change cells & bar chart
  ✔ Severity scoring  – Critical / Major / Minor auto-tagged
  ✔ Configurable via YAML / inline config dataclass
  ✔ Parallel EDB extraction for speed
  ✔ Numeric tolerances – ignore sub-threshold changes
  ✔ Robust error handling with per-check fallbacks
  ✔ [v3] Local Git integration – auto-commit, branch diff, log history
  ✔ [v3] SQLite database – persist every run, query history, trend tracking
  ✔ [v3] EM simulation criticality – smart rules flag changes needing EM re-run
  ✔ [v3] Baseline reference suite – lock golden snapshots, validate regression

Requirements:
    pip install pyedb pandas openpyxl rich pyyaml gitpython

Usage:
    # Basic diff
    python edb_diff_regression.py \\
        --baseline  path/to/baseline.aedb \\
        --modified  path/to/modified.aedb \\
        --report    diff_report.xlsx

    # With Git, DB, and EM-criticality tagging
    python edb_diff_regression.py \\
        --baseline  path/to/baseline.aedb \\
        --modified  path/to/modified.aedb \\
        --git-repo  /path/to/project/repo \\
        --db        edb_history.db \\
        --report    diff_report.xlsx \\
        --simulate

    # Save current design as a named baseline reference
    python edb_diff_regression.py \\
        --save-baseline  golden_v1.2 \\
        --modified       path/to/design.aedb \\
        --db             edb_history.db

    # Validate modified against all stored baseline references
    python edb_diff_regression.py \\
        --validate-baselines \\
        --modified  path/to/design.aedb \\
        --db        edb_history.db \\
        --report    validation_report.xlsx

    # Show run history from DB
    python edb_diff_regression.py --db edb_history.db --history

    [--config   config.yaml]
    [--freq-ghz 20]
    [--json     diff.json]
    [--demo]
"""

# ── stdlib ────────────────────────────────────────────────────────────────────
import argparse
from concurrent.futures import ThreadPoolExecutor, as_completed
import dataclasses
from dataclasses import asdict, dataclass, field
from datetime import datetime
from enum import Enum
import json
import logging
from pathlib import Path
import re as _re
import sqlite3
import sys
import time
import traceback
from typing import Any, Callable, Optional

# ── third-party ───────────────────────────────────────────────────────────────
_MISSING: list = []
try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    import pandas as pd
except ImportError:
    _MISSING.append("pandas openpyxl")

try:
    from rich.console import Console
    from rich.panel import Panel
    from rich.table import Table

    console = Console()
    HAS_RICH = True
except ImportError:
    HAS_RICH = False

    class _FallbackConsole:
        def print(self, *a, **kw):
            print(*a)

        def rule(self, t=""):
            print(f"\n{'─' * 60} {t} {'─' * 60}")

    console = _FallbackConsole()

try:
    import yaml

    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    from pyedb import Edb

    HAS_EDB = True
except ImportError:
    HAS_EDB = False

try:
    import git

    HAS_GIT = True
except ImportError:
    HAS_GIT = False

if _MISSING:
    sys.exit(f"Missing packages – run:  pip install {' '.join(_MISSING)}")

# ── logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("edb_diff")


class ChangeType(str, Enum):
    ADDED = "Added"
    REMOVED = "Removed"
    MODIFIED = "Modified"


class Severity(str, Enum):
    CRITICAL = "Critical"
    MAJOR = "Major"
    MINOR = "Minor"


class Category(str, Enum):
    NET = "Net"
    COMPONENT = "Component"
    LAYER = "Layer"
    PADSTACK = "Padstack"
    TRACE = "Trace"
    VIA = "Via"
    RULE = "Design Rule"


class EMCriticalReason(str, Enum):
    """Reasons why a change is flagged as requiring EM re-simulation."""

    DIELECTRIC_ER = "Dielectric Er changed"
    DIELECTRIC_LOSS_TANGENT = "Loss tangent changed"
    LAYER_THICKNESS = "Layer thickness changed (>5%)"
    STACKUP_REORDER = "Layer order changed"
    TRACE_WIDTH_NARROW = "Trace width narrowed (>10%)"
    TRACE_LAYER_CHANGE = "Trace moved to different layer"
    VIA_DRILL_CHANGE = "Via drill size changed (>15%)"
    VIA_COUNT_DECREASE = "Via count decreased on power/RF net"
    DIFF_PAIR_MODIFIED = "Differential pair geometry changed"
    RF_NET_MODIFIED = "RF/high-speed net geometry changed"
    POWER_NET_TOPOLOGY = "Power net topology changed"
    COMPONENT_MOVE_LARGE = "Component moved significantly (>0.5 mm)"
    NEW_HIGH_SPEED_COMP = "High-speed component added/replaced"


# color palette
_P = {
    "header_bg": "1F3864",
    "header_fg": "FFFFFF",
    "title_bg": "2E75B6",
    "added_bg": "E2EFDA",
    "removed_bg": "FCE4D6",
    "modified_bg": "FFF2CC",
    "critical_bg": "FF0000",
    "major_bg": "FF9900",
    "minor_bg": "FFFF00",
    "pass_bg": "00B050",
    "fail_bg": "C00000",
    "warn_bg": "FF9900",
    "alt_row": "F5F5F5",
    "em_req_bg": "7030A0",  # purple – EM re-simulation required
    "git_bg": "203864",
    "baseline_ok": "00B050",
    "baseline_fail": "C00000",
}


def _fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


@dataclass
class DiffConfig:
    edb_version: str = "2026.1"
    # Numeric tolerances (changes below these are suppressed)
    thickness_tol_um: float = 0.5
    position_tol_mm: float = 0.01
    trace_width_tol_mm: float = 0.001
    drill_tol_mm: float = 0.005
    # Simulation
    sim_start_freq: str = "1MHz"
    sim_stop_freq: str = "10GHz"
    sim_step_freq: str = "500MHz"
    sim_solver: str = "SiwaveSYZ"
    # Regression thresholds
    thresh_IL_dB: float = 2.0
    thresh_RL_dB: float = 3.0
    thresh_impedance_ohm: float = 5.0
    # Severity rules
    critical_categories: list = field(default_factory=lambda: [Category.RULE.value])
    major_categories: list = field(default_factory=lambda: [Category.LAYER.value, Category.PADSTACK.value])
    max_workers: int = 2
    # ── EM criticality thresholds ────────────────────────────────────────
    em_layer_thickness_pct: float = 5.0  # % thickness change triggers EM flag
    em_trace_width_pct: float = 10.0  # % trace width change triggers EM flag
    em_via_drill_pct: float = 15.0  # % drill change triggers EM flag
    em_via_count_pct: float = 20.0  # % via-count decrease triggers EM flag
    em_component_move_mm: float = 0.5  # mm component displacement triggers EM flag
    # Net name patterns that are considered RF/high-speed (regex-compatible)
    em_rf_net_patterns: list = field(
        default_factory=lambda: ["RF", "ANT", "DIFF", "CLK", "DDR", "SERDES", "TX", "RX", "HS_"]
    )
    em_power_net_patterns: list = field(default_factory=lambda: ["VDD", "VCC", "PWR", "GND", "VSS", "VBAT"])
    # High-speed component reference patterns
    em_hs_comp_patterns: list = field(default_factory=lambda: ["U", "IC", "MCU", "FPGA", "DDR", "PHY"])
    git_auto_commit: bool = False
    git_commit_message: str = "edb_diff: automated design checkpoint"
    db_path: str = ""

    @classmethod
    def from_yaml(cls, path: str) -> "DiffConfig":
        if not HAS_YAML:
            log.warning("pyyaml not installed – using default config.")
            return cls()
        with open(path) as f:
            data = yaml.safe_load(f)
        valid_keys = {f.name for f in dataclasses.fields(cls)}
        return cls(**{k: v for k, v in (data or {}).items() if k in valid_keys})

    def to_yaml(self, path: str):
        if not HAS_YAML:
            return
        with open(path, "w") as f:
            yaml.dump(asdict(self), f, default_flow_style=False)


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class DiffEntry:
    category: str
    change_type: str
    item_name: str
    baseline_value: Any = ""
    modified_value: Any = ""
    detail: str = ""
    severity: str = Severity.MINOR.value
    impact_summary: str = ""
    em_sim_required: bool = False
    em_sim_reasons: list = field(default_factory=list)  # list[str]

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class RegressionResult:
    net_name: str
    metric: str
    frequency_ghz: float
    baseline_value: float
    modified_value: float
    delta: float
    threshold: float
    pass_fail: str
    impact: str = ""

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class ExtractionError:
    phase: str
    message: str
    tb: str = ""


@dataclass
class BaselineValidationResult:
    """Result of comparing the modified design against a stored baseline snapshot."""

    baseline_name: str
    baseline_tag: str
    stored_at: str
    total_changes: int
    critical_count: int
    major_count: int
    minor_count: int
    em_required_count: int
    regression_fails: int
    passed: bool
    summary: str

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class GitInfo:
    """Metadata captured from a Git repository at diff time."""

    repo_path: str = ""
    current_branch: str = ""
    current_commit: str = ""
    commit_message: str = ""
    is_dirty: bool = False
    changed_files: list = field(default_factory=list)
    new_commit_hash: str = ""  # set if auto-commit was performed

    def to_dict(self) -> dict:
        return asdict(self)


@dataclass
class DiffReport:
    baseline_path: str
    modified_path: str
    generated_at: str = field(default_factory=lambda: datetime.now().isoformat(timespec="seconds"))
    config: dict = field(default_factory=dict)
    diffs: list = field(default_factory=list)  # list[DiffEntry]
    regression: list = field(default_factory=list)  # list[RegressionResult]
    errors: list = field(default_factory=list)  # list[ExtractionError]
    elapsed_sec: float = 0.0
    # v3
    git_info: Optional[GitInfo] = None
    run_id: int = 0  # populated after DB insert
    em_critical_entries: list = field(default_factory=list)  # subset of diffs
    baseline_validations: list = field(default_factory=list)  # list[BaselineValidationResult]

    @property
    def summary(self) -> dict:
        cats: dict = {}
        for d in self.diffs:
            key = f"{d.category} – {d.change_type}"
            cats[key] = cats.get(key, 0) + 1
        return dict(sorted(cats.items()))

    @property
    def severity_counts(self) -> dict:
        counts = {s.value: 0 for s in Severity}
        for d in self.diffs:
            counts[d.severity] = counts.get(d.severity, 0) + 1
        return counts

    @property
    def regression_summary(self) -> dict:
        counts = {"PASS": 0, "WARN": 0, "FAIL": 0}
        for r in self.regression:
            counts[r.pass_fail] = counts.get(r.pass_fail, 0) + 1
        return counts

    @property
    def em_required_count(self) -> int:
        return sum(1 for d in self.diffs if d.em_sim_required)

    def to_dict(self) -> dict:
        return {
            "run_id": self.run_id,
            "baseline_path": self.baseline_path,
            "modified_path": self.modified_path,
            "generated_at": self.generated_at,
            "elapsed_sec": self.elapsed_sec,
            "config": self.config,
            "summary": self.summary,
            "severity_counts": self.severity_counts,
            "regression_summary": self.regression_summary,
            "em_required_count": self.em_required_count,
            "diffs": [d.to_dict() for d in self.diffs],
            "regression": [r.to_dict() for r in self.regression],
            "errors": [asdict(e) for e in self.errors],
            "git_info": self.git_info.to_dict() if self.git_info else None,
            "baseline_validations": [b.to_dict() for b in self.baseline_validations],
        }


# ─────────────────────────────────────────────────────────────────────────────
# Git Integration
# ─────────────────────────────────────────────────────────────────────────────


def _capture_git_info(repo_path: str, auto_commit: bool = False, commit_msg: str = "") -> GitInfo:
    """Capture current Git state; optionally auto-commit dirty working tree."""
    if not HAS_GIT:
        log.warning("gitpython not installed – Git integration disabled. Install with: pip install gitpython")
        return GitInfo(repo_path=repo_path)
    try:
        repo = git.Repo(repo_path, search_parent_directories=True)
        head = repo.head.commit
        info = GitInfo(
            repo_path=repo.working_dir,
            current_branch=repo.active_branch.name if not repo.head.is_detached else "DETACHED",
            current_commit=head.hexsha[:12],
            commit_message=head.message.strip(),
            is_dirty=repo.is_dirty(untracked_files=True),
            changed_files=[item.a_path for item in repo.index.diff(None)] + repo.untracked_files,
        )
        if auto_commit and repo.is_dirty(untracked_files=False):
            repo.git.add(A=True)
            new_commit = repo.index.commit(commit_msg or "edb_diff: automated checkpoint")
            info.new_commit_hash = new_commit.hexsha[:12]
            log.info(f"[Git] Auto-committed: {info.new_commit_hash}")
        return info
    except Exception as exc:
        log.warning(f"[Git] Could not read repo at {repo_path}: {exc}")
        return GitInfo(repo_path=repo_path)


def get_git_log(repo_path: str, max_entries: int = 20) -> list[dict]:
    """Return recent commits as a list of dicts."""
    if not HAS_GIT:
        return []
    try:
        repo = git.Repo(repo_path, search_parent_directories=True)
        entries = []
        for commit in list(repo.iter_commits())[:max_entries]:
            entries.append(
                {
                    "hash": commit.hexsha[:12],
                    "author": str(commit.author),
                    "date": datetime.fromtimestamp(commit.committed_date).isoformat(),
                    "message": commit.message.strip()[:120],
                }
            )
        return entries
    except Exception as exc:
        log.warning(f"[Git] Log failed: {exc}")
        return []


# ─────────────────────────────────────────────────────────────────────────────
# SQLite Database
# ─────────────────────────────────────────────────────────────────────────────

_DB_SCHEMA = """
CREATE TABLE IF NOT EXISTS runs (
    id               INTEGER PRIMARY KEY AUTOINCREMENT,
    generated_at     TEXT NOT NULL,
    baseline_path    TEXT NOT NULL,
    modified_path    TEXT NOT NULL,
    elapsed_sec      REAL,
    total_diffs      INTEGER,
    critical_count   INTEGER,
    major_count      INTEGER,
    minor_count      INTEGER,
    em_required      INTEGER,
    reg_pass         INTEGER,
    reg_warn         INTEGER,
    reg_fail         INTEGER,
    git_branch       TEXT,
    git_commit       TEXT,
    git_dirty        INTEGER,
    notes            TEXT,
    full_json        TEXT
);

CREATE TABLE IF NOT EXISTS diffs (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id          INTEGER NOT NULL REFERENCES runs(id),
    category        TEXT,
    change_type     TEXT,
    item_name       TEXT,
    baseline_value  TEXT,
    modified_value  TEXT,
    detail          TEXT,
    severity        TEXT,
    impact_summary  TEXT,
    em_sim_required INTEGER,
    em_sim_reasons  TEXT
);

CREATE TABLE IF NOT EXISTS regression_results (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id          INTEGER NOT NULL REFERENCES runs(id),
    net_name        TEXT,
    metric          TEXT,
    frequency_ghz   REAL,
    baseline_value  REAL,
    modified_value  REAL,
    delta           REAL,
    threshold       REAL,
    pass_fail       TEXT,
    impact          TEXT
);

CREATE TABLE IF NOT EXISTS baseline_references (
    id              INTEGER PRIMARY KEY AUTOINCREMENT,
    name            TEXT UNIQUE NOT NULL,
    tag             TEXT,
    stored_at       TEXT NOT NULL,
    design_path     TEXT,
    git_commit      TEXT,
    git_branch      TEXT,
    snapshot_json   TEXT NOT NULL,
    notes           TEXT
);

CREATE TABLE IF NOT EXISTS baseline_validations (
    id                  INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id              INTEGER NOT NULL REFERENCES runs(id),
    baseline_ref_id     INTEGER REFERENCES baseline_references(id),
    baseline_name       TEXT,
    total_changes       INTEGER,
    critical_count      INTEGER,
    major_count         INTEGER,
    minor_count         INTEGER,
    em_required_count   INTEGER,
    regression_fails    INTEGER,
    passed              INTEGER,
    summary             TEXT
);
"""


class DiffDatabase:
    """SQLite-backed persistence layer for diff runs, diffs, and baseline references."""

    def __init__(self, db_path: str):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row
        self._init_schema()
        log.info(f"[DB] Connected to {db_path}")

    def _init_schema(self):
        self.conn.executescript(_DB_SCHEMA)
        self.conn.commit()
        self._migrate_schema()

    def _migrate_schema(self):
        """Apply incremental migrations so existing databases gain new columns."""
        migrations = [
            # minor_count added to baseline_validations
            "ALTER TABLE baseline_validations ADD COLUMN minor_count INTEGER",
            # notes column on runs (was in schema but may be absent in old DBs)
            "ALTER TABLE runs ADD COLUMN notes TEXT",
        ]
        for sql in migrations:
            try:
                self.conn.execute(sql)
                self.conn.commit()
            except sqlite3.OperationalError:
                pass  # Column already exists – safe to ignore

    def save_run(self, report: DiffReport) -> int:
        """Insert a completed DiffReport and return the assigned run_id."""
        sc = report.severity_counts
        rs = report.regression_summary
        gi = report.git_info or GitInfo()
        cur = self.conn.execute(
            """INSERT INTO runs
               (generated_at, baseline_path, modified_path, elapsed_sec,
                total_diffs, critical_count, major_count, minor_count,
                em_required, reg_pass, reg_warn, reg_fail,
                git_branch, git_commit, git_dirty, full_json)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (
                report.generated_at,
                report.baseline_path,
                report.modified_path,
                report.elapsed_sec,
                len(report.diffs),
                sc.get(Severity.CRITICAL.value, 0),
                sc.get(Severity.MAJOR.value, 0),
                sc.get(Severity.MINOR.value, 0),
                report.em_required_count,
                rs.get("PASS", 0),
                rs.get("WARN", 0),
                rs.get("FAIL", 0),
                gi.current_branch,
                gi.current_commit,
                int(gi.is_dirty),
                json.dumps(report.to_dict(), default=str),
            ),
        )
        run_id = cur.lastrowid
        # Insert individual diffs
        for d in report.diffs:
            self.conn.execute(
                """INSERT INTO diffs
                   (run_id, category, change_type, item_name, baseline_value,
                    modified_value, detail, severity, impact_summary,
                    em_sim_required, em_sim_reasons)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    run_id,
                    d.category,
                    d.change_type,
                    d.item_name,
                    str(d.baseline_value),
                    str(d.modified_value),
                    d.detail,
                    d.severity,
                    d.impact_summary,
                    int(d.em_sim_required),
                    json.dumps(d.em_sim_reasons),
                ),
            )
        # Insert regression results
        for r in report.regression:
            self.conn.execute(
                """INSERT INTO regression_results
                   (run_id, net_name, metric, frequency_ghz, baseline_value,
                    modified_value, delta, threshold, pass_fail, impact)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    run_id,
                    r.net_name,
                    r.metric,
                    r.frequency_ghz,
                    r.baseline_value,
                    r.modified_value,
                    r.delta,
                    r.threshold,
                    r.pass_fail,
                    r.impact,
                ),
            )
        # Insert baseline validation results
        for bv in report.baseline_validations:
            self.conn.execute(
                """INSERT INTO baseline_validations
                   (run_id, baseline_name, total_changes, critical_count,
                    major_count, minor_count, em_required_count, regression_fails, passed, summary)
                   VALUES (?,?,?,?,?,?,?,?,?,?)""",
                (
                    run_id,
                    bv.baseline_name,
                    bv.total_changes,
                    bv.critical_count,
                    bv.major_count,
                    bv.minor_count,
                    bv.em_required_count,
                    bv.regression_fails,
                    int(bv.passed),
                    bv.summary,
                ),
            )
        self.conn.commit()
        log.info(f"[DB] Run saved – run_id={run_id}")
        return run_id

    def save_baseline_reference(
        self, name: str, tag: str, design_path: str, snapshot: dict, git_info: Optional[GitInfo] = None, notes: str = ""
    ) -> int:
        """Store an EDB snapshot as a named baseline reference."""
        gi = git_info or GitInfo()
        try:
            cur = self.conn.execute(
                """INSERT OR REPLACE INTO baseline_references
                   (name, tag, stored_at, design_path, git_commit, git_branch,
                    snapshot_json, notes)
                   VALUES (?,?,?,?,?,?,?,?)""",
                (
                    name,
                    tag,
                    datetime.now().isoformat(timespec="seconds"),
                    design_path,
                    gi.current_commit,
                    gi.current_branch,
                    json.dumps(snapshot, default=str),
                    notes,
                ),
            )
            self.conn.commit()
            log.info(f"[DB] Baseline reference '{name}' saved (id={cur.lastrowid})")
            return cur.lastrowid
        except Exception as exc:
            log.error(f"[DB] save_baseline_reference failed: {exc}")
            raise

    def load_baseline_reference(self, name: str) -> Optional[dict]:
        """Load a named baseline reference snapshot dict."""
        row = self.conn.execute("SELECT * FROM baseline_references WHERE name=?", (name,)).fetchone()
        if row is None:
            return None
        d = dict(row)
        d["snapshot"] = json.loads(d["snapshot_json"])
        return d

    def list_baseline_references(self) -> list[dict]:
        rows = self.conn.execute(
            "SELECT id, name, tag, stored_at, design_path, git_commit, git_branch, notes "
            "FROM baseline_references ORDER BY stored_at DESC"
        ).fetchall()
        return [dict(r) for r in rows]

    def get_run_history(self, limit: int = 50) -> list[dict]:
        rows = self.conn.execute(
            """SELECT id, generated_at, baseline_path, modified_path,
                      total_diffs, critical_count, major_count, minor_count,
                      em_required, reg_pass, reg_warn, reg_fail,
                      git_branch, git_commit, git_dirty
               FROM runs ORDER BY id DESC LIMIT ?""",
            (limit,),
        ).fetchall()
        return [dict(r) for r in rows]

    _TREND_FIELDS = frozenset(
        {
            "total_diffs",
            "critical_count",
            "major_count",
            "minor_count",
            "em_required",
            "reg_pass",
            "reg_warn",
            "reg_fail",
            "elapsed_sec",
        }
    )

    def get_trend(self, field_name: str = "critical_count", last_n: int = 20) -> list[dict]:
        """Return (generated_at, value) pairs for trend analysis.

        ``field_name`` is validated against an allowlist to prevent SQL injection.
        """
        if field_name not in self._TREND_FIELDS:
            raise ValueError(f"Invalid trend field '{field_name}'. Must be one of: {sorted(self._TREND_FIELDS)}")
        rows = self.conn.execute(
            f"SELECT generated_at, {field_name} FROM runs ORDER BY id DESC LIMIT ?",  # nosec – allowlisted
            (last_n,),
        ).fetchall()
        return [dict(r) for r in reversed(rows)]

    def close(self):
        self.conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# EM Simulation Criticality Engine
# ─────────────────────────────────────────────────────────────────────────────


def _net_matches(net_name: str, patterns: list[str]) -> bool:
    name_upper = net_name.upper()
    return any(_re.search(p.upper(), name_upper) for p in patterns)


def _pct_change(old: Any, new: Any) -> float:
    """Return absolute % change, or 0 if non-numeric."""
    try:
        old_f, new_f = float(old), float(new)
        if old_f == 0:
            return 0.0
        return abs(new_f - old_f) / abs(old_f) * 100.0
    except (TypeError, ValueError):
        return 0.0


def evaluate_em_criticality(entry: DiffEntry, cfg: DiffConfig) -> DiffEntry:
    """
    Analyze a DiffEntry and tag it with em_sim_required=True / em_sim_reasons
    if it satisfies any EM criticality rule.  Returns the (mutated) entry.
    """
    reasons: list[str] = []
    cat = entry.category
    ct = entry.change_type
    det = entry.detail
    bv = entry.baseline_value
    mv = entry.modified_value
    name = entry.item_name

    # ── Layer stackup rules ──────────────────────────────────────────────────
    if cat == Category.LAYER.value:
        if ct == ChangeType.MODIFIED.value:
            if det == "Er":
                reasons.append(EMCriticalReason.DIELECTRIC_ER.value)
            elif det == "loss_tangent":
                reasons.append(EMCriticalReason.DIELECTRIC_LOSS_TANGENT.value)
            elif det == "thickness_um":
                if _pct_change(bv, mv) > cfg.em_layer_thickness_pct:
                    reasons.append(EMCriticalReason.LAYER_THICKNESS.value)
            elif det == "stack_order":
                reasons.append(EMCriticalReason.STACKUP_REORDER.value)

    # ── Trace rules ──────────────────────────────────────────────────────────
    elif cat == Category.TRACE.value:
        if ct == ChangeType.MODIFIED.value:
            if det in ("min_width_mm", "max_width_mm"):
                # Flag narrowing only
                try:
                    if float(mv) < float(bv):
                        if _pct_change(bv, mv) > cfg.em_trace_width_pct:
                            reasons.append(EMCriticalReason.TRACE_WIDTH_NARROW.value)
                except (TypeError, ValueError):
                    pass
            elif det == "layers":
                reasons.append(EMCriticalReason.TRACE_LAYER_CHANGE.value)
        if _net_matches(name, cfg.em_rf_net_patterns):
            reasons.append(EMCriticalReason.RF_NET_MODIFIED.value)

    # ── Via / Padstack rules ─────────────────────────────────────────────────
    elif cat == Category.PADSTACK.value:
        if ct == ChangeType.MODIFIED.value and det == "drill_mm":
            if _pct_change(bv, mv) > cfg.em_via_drill_pct:
                reasons.append(EMCriticalReason.VIA_DRILL_CHANGE.value)

    elif cat == Category.VIA.value:
        if ct == ChangeType.MODIFIED.value and det == "via_count":
            try:
                if float(mv) < float(bv):  # decrease only
                    if _pct_change(bv, mv) > cfg.em_via_count_pct:
                        if _net_matches(name, cfg.em_power_net_patterns + cfg.em_rf_net_patterns):
                            reasons.append(EMCriticalReason.VIA_COUNT_DECREASE.value)
            except (TypeError, ValueError):
                pass

    # ── Net topology rules ───────────────────────────────────────────────────
    elif cat == Category.NET.value:
        if _net_matches(name, cfg.em_power_net_patterns) and ct != ChangeType.MODIFIED.value:
            reasons.append(EMCriticalReason.POWER_NET_TOPOLOGY.value)

    # ── Component rules ──────────────────────────────────────────────────────
    elif cat == Category.COMPONENT.value:
        if ct == ChangeType.MODIFIED.value:
            if det in ("center_x_mm", "center_y_mm"):
                try:
                    if abs(float(mv) - float(bv)) > cfg.em_component_move_mm:
                        reasons.append(EMCriticalReason.COMPONENT_MOVE_LARGE.value)
                except (TypeError, ValueError):
                    pass
            if det == "part_name":
                # Any IC/high-speed component type change
                if _net_matches(name, cfg.em_hs_comp_patterns) or any(
                    name.upper().startswith(p.upper()) for p in cfg.em_hs_comp_patterns
                ):
                    reasons.append(EMCriticalReason.NEW_HIGH_SPEED_COMP.value)
        elif ct in (ChangeType.ADDED.value, ChangeType.REMOVED.value):
            if any(name.upper().startswith(p.upper()) for p in cfg.em_hs_comp_patterns):
                reasons.append(EMCriticalReason.NEW_HIGH_SPEED_COMP.value)

    # ── Deduplicate and promote severity ────────────────────────────────────
    entry.em_sim_reasons = list(dict.fromkeys(reasons))
    entry.em_sim_required = bool(reasons)
    if entry.em_sim_required and entry.severity != Severity.CRITICAL.value:
        entry.severity = Severity.CRITICAL.value  # EM critical → Critical severity
    return entry


def apply_em_criticality(diffs: list, cfg: DiffConfig) -> list:
    """Tag all DiffEntries in-place; return list of EM-critical entries."""
    em_entries = []
    for entry in diffs:
        evaluate_em_criticality(entry, cfg)
        if entry.em_sim_required:
            em_entries.append(entry)
    return em_entries


# ─────────────────────────────────────────────────────────────────────────────
# Baseline Reference management
# ─────────────────────────────────────────────────────────────────────────────


def save_baseline_reference(
    db: DiffDatabase,
    name: str,
    design_path: str,
    cfg: Optional["DiffConfig"] = None,
    git_info: Optional[GitInfo] = None,
    tag: str = "",
    notes: str = "",
) -> int:
    """
    Open an EDB, extract its full snapshot, and persist it as a named
    baseline reference in the database.
    """
    cfg = cfg or DiffConfig()
    if HAS_EDB:
        log.info(f"[Baseline] Extracting snapshot from {design_path} …")
        edb = _open_edb(design_path, cfg.edb_version)
        try:
            snapshot = extract_all(edb)
        finally:
            closer = getattr(edb, "close", None) or getattr(edb, "close_edb", None)
            if closer:
                closer()
    else:
        log.warning("[Baseline] pyedb not available – storing empty snapshot.")
        snapshot = {}
    ref_id = db.save_baseline_reference(
        name=name, tag=tag, design_path=design_path, snapshot=snapshot, git_info=git_info, notes=notes
    )
    console.print(f"[bold green]✔ Baseline reference '{name}' saved (id={ref_id})[/bold green]")
    return ref_id


def validate_against_baselines(
    modified_path: str,
    db: DiffDatabase,
    cfg: Optional[DiffConfig] = None,
    baseline_names: Optional[list[str]] = None,
) -> tuple[list[BaselineValidationResult], list[DiffReport]]:
    """
    Compare the modified design against all (or named) stored baselines.
    Returns (validation_results, per-baseline diff reports).
    """
    cfg = cfg or DiffConfig()
    refs = db.list_baseline_references()
    if baseline_names:
        refs = [r for r in refs if r["name"] in baseline_names]
    if not refs:
        log.warning("[Baseline] No baseline references found in DB.")
        return [], []

    validations: list[BaselineValidationResult] = []
    sub_reports: list[DiffReport] = []

    # Extract modified design once
    if HAS_EDB:
        edb_mod = _open_edb(modified_path, cfg.edb_version)
        try:
            mod_data = extract_all(edb_mod)
        finally:
            closer = getattr(edb_mod, "close", None) or getattr(edb_mod, "close_edb", None)
            if closer:
                closer()
    else:
        mod_data = {}

    for ref in refs:
        ref_data = db.load_baseline_reference(ref["name"])
        if ref_data is None:
            continue
        baseline_snapshot = ref_data["snapshot"]

        sub_report = DiffReport(
            baseline_path=ref["design_path"] or ref["name"],
            modified_path=modified_path,
        )
        for key, (cat, skip) in _EXTRACTOR_CATEGORIES.items():
            try:
                entries = _dict_diff(
                    cat.value,
                    baseline_snapshot.get(key, {}),
                    mod_data.get(key, {}),
                    cfg,
                    skip_fields=skip,
                )
                sub_report.diffs.extend(entries)
            except Exception as exc:
                sub_report.errors.append(
                    ExtractionError(phase=f"validate:{key}", message=str(exc), tb=traceback.format_exc())
                )

        em_critical = apply_em_criticality(sub_report.diffs, cfg)
        sub_report.em_critical_entries = em_critical
        sc = sub_report.severity_counts

        passed = sc.get(Severity.CRITICAL.value, 0) == 0 and len(em_critical) == 0
        summary_parts = [f"{len(sub_report.diffs)} total change(s)"]
        if sc[Severity.CRITICAL.value]:
            summary_parts.append(f"{sc[Severity.CRITICAL.value]} critical")
        if em_critical:
            summary_parts.append(f"{len(em_critical)} EM-required")
        summary_msg = "PASS – " if passed else "FAIL – "
        summary_msg += ", ".join(summary_parts)

        vr = BaselineValidationResult(
            baseline_name=ref["name"],
            baseline_tag=ref.get("tag", ""),
            stored_at=ref["stored_at"],
            total_changes=len(sub_report.diffs),
            critical_count=sc.get(Severity.CRITICAL.value, 0),
            major_count=sc.get(Severity.MAJOR.value, 0),
            minor_count=sc.get(Severity.MINOR.value, 0),
            em_required_count=len(em_critical),
            regression_fails=0,
            passed=passed,
            summary=summary_msg,
        )
        validations.append(vr)
        sub_reports.append(sub_report)

    return validations, sub_reports


# ─────────────────────────────────────────────────────────────────────────────
# EDB helpers
# ─────────────────────────────────────────────────────────────────────────────


def _safe(fn: Callable, default=None, label: str = ""):
    try:
        return fn()
    except Exception as exc:
        log.debug(f"[{label}] {exc}")
        return default


def _open_edb(path: str, version: str) -> "Edb":
    if not HAS_EDB:
        raise RuntimeError("pyedb is not installed.")
    log.info(f"Opening EDB: {path}")
    return Edb(edbpath=path, edbversion=version)


# ── Extractors ────────────────────────────────────────────────────────────────


def extract_net_info(edb: "Edb") -> dict:
    info: dict = {}
    for name, net in edb.nets.nets.items():
        prims = _safe(lambda n=net: n.primitives, default=[], label="net.primitives")
        info[name] = {
            "component_count": _safe(lambda n=net: len(n.components), default=0),
            "primitive_count": len(prims) if prims else 0,
            "is_power": _safe(lambda n=net: n.is_power_ground, default=False),
        }
    return info


def extract_component_info(edb: "Edb") -> dict:
    info: dict = {}
    for ref, comp in edb.components.components.items():
        bb = _safe(lambda c=comp: c.bounding_box, default=[0, 0, 0, 0])
        cx = round((bb[0] + bb[2]) / 2 * 1e3, 4) if bb else 0.0
        cy = round((bb[1] + bb[3]) / 2 * 1e3, 4) if bb else 0.0
        info[ref] = {
            "part_name": _safe(lambda c=comp: c.partname, default=""),
            "placement_layer": _safe(lambda c=comp: c.placement_layer, default=""),
            "center_x_mm": cx,
            "center_y_mm": cy,
            "comp_type": str(_safe(lambda c=comp: c.type, default="")),
            "value": _safe(lambda c=comp: c.value, default=""),
            "num_pins": _safe(lambda c=comp: len(c.pins), default=0),
        }
    return info


def extract_stackup_info(edb: "Edb") -> dict:
    info: dict = {}
    for idx, (name, layer) in enumerate(edb.stackup.stackup_layers.items()):
        mat = _safe(lambda ly=layer: ly.material, default="")
        mat_props: dict = {}
        if mat:
            try:
                mo = edb.materials[mat]
                mat_props = {
                    "Er": round(_safe(lambda m=mo: m.permittivity, default=0.0), 4),
                    "loss_tangent": round(_safe(lambda m=mo: m.loss_tangent, default=0.0), 6),
                    "conductivity": round(_safe(lambda m=mo: m.conductivity, default=0.0), 2),
                }
            except Exception:
                pass
        info[name] = {
            "stack_order": idx,
            "layer_type": str(_safe(lambda ly=layer: ly.type, default="")),
            "thickness_um": round(_safe(lambda ly=layer: ly.thickness, default=0.0) * 1e6, 3),
            "material": mat,
            **mat_props,
            "dielectric_fill": _safe(lambda ly=layer: ly.dielectric_fill, default=""),
        }
    return info


def extract_padstack_info(edb: "Edb") -> dict:
    info: dict = {}
    for name, ps in edb.padstacks.definitions.items():
        layers = _safe(lambda p=ps: p.pad_by_layer, default={})
        per_layer: dict = {}
        for lname, pad in (layers or {}).items():
            try:
                pv = pad.parameters_values
                per_layer[lname] = {
                    "pad_shape": str(pad.geometry_type),
                    "pad_x_mm": round(pv[0] * 1e3, 4) if pv else 0,
                    "pad_y_mm": round(pv[1] * 1e3, 4) if len(pv) > 1 else 0,
                }
            except Exception:
                per_layer[lname] = {}
        info[name] = {
            "drill_mm": round((_safe(lambda p=ps: p.hole_range) or 0) * 1e3, 4),
            "antipad_mm": round((_safe(lambda p=ps: p.antipad_hole_range) or 0) * 1e3, 4),
            "via_material": _safe(lambda p=ps: p.via_material, default=""),
            "plating_ratio": _safe(lambda p=ps: p.hole_plating_ratio, default=0),
            "pad_layers": per_layer,
        }
    return info


def extract_trace_info(edb: "Edb") -> dict:
    info: dict = {}
    try:
        for name, net in edb.nets.nets.items():
            prims = _safe(lambda n=net: n.primitives, default=[]) or []
            traces = [p for p in prims if _safe(lambda p=p: p.type, default="") == "path"]
            widths, lengths, lyrs = [], [], set()
            for t in traces:
                w = _safe(lambda t=t: t.width)
                if w is not None:
                    widths.append(round(w * 1e3, 4))
                ln = _safe(lambda t=t: t.length)
                if ln is not None:
                    lengths.append(round(ln * 1e3, 4))
                ly = _safe(lambda t=t: t.layer_name)
                if ly:
                    lyrs.add(ly)
            info[name] = {
                "trace_count": len(traces),
                "min_width_mm": min(widths, default=0),
                "max_width_mm": max(widths, default=0),
                "total_length_mm": round(sum(lengths), 4),
                "layers": sorted(lyrs),
            }
    except Exception as exc:
        log.warning(f"Trace extraction failed: {exc}")
    return info


def extract_via_info(edb: "Edb") -> dict:
    info: dict = {}
    try:
        for name, net in edb.nets.nets.items():
            prims = _safe(lambda n=net: n.primitives, default=[]) or []
            vias = [p for p in prims if _safe(lambda p=p: p.type, default="") == "Via"]
            info[name] = {"via_count": len(vias)}
    except Exception as exc:
        log.warning(f"Via extraction failed: {exc}")
    return info


def extract_all(edb: "Edb") -> dict:
    return {
        "nets": extract_net_info(edb),
        "components": extract_component_info(edb),
        "stackup": extract_stackup_info(edb),
        "padstacks": extract_padstack_info(edb),
        "traces": extract_trace_info(edb),
        "vias": extract_via_info(edb),
    }


# ─────────────────────────────────────────────────────────────────────────────
# Diff engine
# ─────────────────────────────────────────────────────────────────────────────


def _auto_severity(category: str, change_type: str, cfg: DiffConfig) -> Severity:
    if category in cfg.critical_categories:
        return Severity.CRITICAL
    if category in cfg.major_categories:
        return Severity.MAJOR
    if change_type in (ChangeType.ADDED.value, ChangeType.REMOVED.value):
        return Severity.MAJOR
    return Severity.MINOR


def _numeric_exceeds(b: Any, m: Any, tol: float) -> bool:
    try:
        return abs(float(m) - float(b)) > tol
    except (TypeError, ValueError):
        return b != m


_TOLERANCE_MAP: dict = {
    (Category.LAYER.value, "thickness_um"): "thickness_tol_um",
    (Category.COMPONENT.value, "center_x_mm"): "position_tol_mm",
    (Category.COMPONENT.value, "center_y_mm"): "position_tol_mm",
    (Category.PADSTACK.value, "drill_mm"): "drill_tol_mm",
    (Category.TRACE.value, "min_width_mm"): "trace_width_tol_mm",
    (Category.TRACE.value, "max_width_mm"): "trace_width_tol_mm",
}


def _tol(category: str, field_name: str, cfg: DiffConfig) -> float:
    attr = _TOLERANCE_MAP.get((category, field_name))
    return getattr(cfg, attr, 0.0) if attr else 0.0


def _impact(category: str, change_type: str, name: str, detail: str, bv: Any, mv: Any) -> str:
    if change_type == ChangeType.ADDED.value:
        return f"New {category.lower()} '{name}' introduced."
    if change_type == ChangeType.REMOVED.value:
        return f"{category} '{name}' removed."
    try:
        delta = round(float(mv) - float(bv), 4)
        pct = round(delta / float(bv) * 100, 1) if float(bv) else 0
        field_label = f"'{detail}'" if detail else "Value"
        return f"{field_label} changed by {delta:+g} ({pct:+.1f}%)."
    except Exception:
        return f"'{detail}' changed: '{bv}' → '{mv}'."


def _dict_diff(
    category: str,
    baseline: dict,
    modified: dict,
    cfg: DiffConfig,
    skip_fields: Optional[set] = None,
) -> list:
    skip_fields = skip_fields or set()
    entries: list = []

    base_keys = set(baseline)
    mod_keys = set(modified)

    for name in sorted(base_keys - mod_keys):
        sev = _auto_severity(category, ChangeType.REMOVED.value, cfg)
        entries.append(
            DiffEntry(
                category=category,
                change_type=ChangeType.REMOVED.value,
                item_name=name,
                baseline_value=str(baseline[name]),
                modified_value="—",
                severity=sev.value,
                impact_summary=_impact(category, ChangeType.REMOVED.value, name, "", "", ""),
            )
        )

    for name in sorted(mod_keys - base_keys):
        sev = _auto_severity(category, ChangeType.ADDED.value, cfg)
        entries.append(
            DiffEntry(
                category=category,
                change_type=ChangeType.ADDED.value,
                item_name=name,
                baseline_value="—",
                modified_value=str(modified[name]),
                severity=sev.value,
                impact_summary=_impact(category, ChangeType.ADDED.value, name, "", "", ""),
            )
        )

    for name in sorted(base_keys & mod_keys):
        bv, mv = baseline[name], modified[name]
        if isinstance(bv, dict) and isinstance(mv, dict):
            for ck in sorted((set(bv) | set(mv)) - skip_fields):
                b, m = bv.get(ck), mv.get(ck)
                if b == m:
                    continue
                t = _tol(category, ck, cfg)
                if t and not _numeric_exceeds(b, m, t):
                    continue
                sev = _auto_severity(category, ChangeType.MODIFIED.value, cfg)
                entries.append(
                    DiffEntry(
                        category=category,
                        change_type=ChangeType.MODIFIED.value,
                        item_name=name,
                        baseline_value=b,
                        modified_value=m,
                        detail=ck,
                        severity=sev.value,
                        impact_summary=_impact(category, ChangeType.MODIFIED.value, name, ck, b, m),
                    )
                )
        elif bv != mv:
            sev = _auto_severity(category, ChangeType.MODIFIED.value, cfg)
            entries.append(
                DiffEntry(
                    category=category,
                    change_type=ChangeType.MODIFIED.value,
                    item_name=name,
                    baseline_value=bv,
                    modified_value=mv,
                    severity=sev.value,
                    impact_summary=_impact(category, ChangeType.MODIFIED.value, name, "", bv, mv),
                )
            )
    return entries


# ─────────────────────────────────────────────────────────────────────────────
# Parallel extraction + orchestration
# ─────────────────────────────────────────────────────────────────────────────

_EXTRACTOR_CATEGORIES = {
    "nets": (Category.NET, set()),
    "components": (Category.COMPONENT, {"layers"}),
    "stackup": (Category.LAYER, set()),
    "padstacks": (Category.PADSTACK, {"pad_layers"}),
    "traces": (Category.TRACE, {"layers"}),
    "vias": (Category.VIA, set()),
}


def _parallel_extract(paths: list, cfg: DiffConfig) -> list:
    labels = ["baseline", "modified"]
    results: list = [None, None]

    def worker(idx: int, path: str):
        label = labels[idx] if idx < len(labels) else str(idx)
        try:
            edb = _open_edb(path, cfg.edb_version)
        except Exception as exc:
            raise RuntimeError(f"Failed to open {label} EDB at '{path}': {exc}") from exc
        try:
            data = extract_all(edb)
        except Exception as exc:
            raise RuntimeError(f"Failed to extract {label} EDB at '{path}': {exc}") from exc
        finally:
            # pyedb exposes close() as the public API; fall back to close_edb()
            closer = getattr(edb, "close", None) or getattr(edb, "close_edb", None)
            if closer:
                try:
                    closer()
                except Exception:
                    pass
        return idx, data

    with ThreadPoolExecutor(max_workers=cfg.max_workers) as pool:
        futures = {pool.submit(worker, i, p): i for i, p in enumerate(paths)}
        for fut in as_completed(futures):
            idx, data = fut.result()  # re-raises any worker exception with clear message
            results[idx] = data

    # Sanity-check: both slots must be populated
    for i, (res, lbl) in enumerate(zip(results, labels)):
        if res is None:
            raise RuntimeError(f"Extraction returned no data for {lbl} (index {i}).")

    return results


def run_diff(
    baseline_path: str,
    modified_path: str,
    cfg: Optional[DiffConfig] = None,
    db: Optional[DiffDatabase] = None,
    git_info: Optional[GitInfo] = None,
) -> DiffReport:
    cfg = cfg or DiffConfig()
    t0 = time.perf_counter()
    report = DiffReport(
        baseline_path=baseline_path,
        modified_path=modified_path,
        config=asdict(cfg),
        git_info=git_info,
    )

    console.rule("Extracting EDB data (parallel)")
    try:
        base_data, mod_data = _parallel_extract([baseline_path, modified_path], cfg)
    except Exception as exc:
        report.errors.append(ExtractionError(phase="extraction", message=str(exc), tb=traceback.format_exc()))
        log.error(f"Extraction failed: {exc}")
        report.elapsed_sec = round(time.perf_counter() - t0, 2)
        return report

    console.rule("Computing diffs")
    for key, (cat, skip) in _EXTRACTOR_CATEGORIES.items():
        try:
            entries = _dict_diff(
                cat.value,
                base_data.get(key, {}),
                mod_data.get(key, {}),
                cfg,
                skip_fields=skip,
            )
            report.diffs.extend(entries)
            log.info(f"  {cat.value}: {len(entries)} change(s)")
        except Exception as exc:
            report.errors.append(ExtractionError(phase=f"diff:{key}", message=str(exc), tb=traceback.format_exc()))
            log.warning(f"  {cat.value} diff failed: {exc}")

    # apply EM criticality tagging
    console.rule("Evaluating EM simulation criticality")
    report.em_critical_entries = apply_em_criticality(report.diffs, cfg)
    log.info(f"  EM re-simulation required: {len(report.em_critical_entries)} change(s)")

    report.elapsed_sec = round(time.perf_counter() - t0, 2)

    # persist to DB
    if db is not None:
        try:
            report.run_id = db.save_run(report)
        except Exception as exc:
            log.warning(f"[DB] Failed to save run: {exc}")

    return report


# ─────────────────────────────────────────────────────────────────────────────
# S-parameter regression
# ─────────────────────────────────────────────────────────────────────────────


def run_simulation_regression(
    baseline_path: str,
    modified_path: str,
    cfg: Optional[DiffConfig] = None,
) -> list:
    cfg = cfg or DiffConfig()

    def _simulate(path: str, label: str) -> dict:
        edb = _open_edb(path, cfg.edb_version)
        try:
            sim = edb.new_simulation_configuration()
            sim.solver_type = getattr(sim.SolverType, cfg.sim_solver, sim.SolverType.SiwaveSYZ)
            sim.start_freq = cfg.sim_start_freq
            sim.stop_freq = cfg.sim_stop_freq
            sim.step_freq = cfg.sim_step_freq
            sim.do_cutout_subdesign = True
            sim.output_aedb = str(Path(path).parent / f"sim_{label}.aedb")
            log.info(f"  [{label}] Launching {cfg.sim_solver} …")
            edb.build_simulation_project(sim)
            sparams = edb.get_statistics()
        finally:
            closer = getattr(edb, "close", None) or getattr(edb, "close_edb", None)
            if closer:
                closer()
        return sparams

    results: list = []
    thresholds = {
        "IL_dB": cfg.thresh_IL_dB,
        "RL_dB": cfg.thresh_RL_dB,
        "Impedance_Ohm": cfg.thresh_impedance_ohm,
    }
    base_sp = _simulate(baseline_path, "baseline")
    mod_sp = _simulate(modified_path, "modified")

    for net in sorted(set(base_sp) & set(mod_sp)):
        for freq in sorted(set(base_sp[net]) & set(mod_sp[net])):
            for metric, thr in thresholds.items():
                bv = base_sp[net][freq].get(metric, 0.0)
                mv = mod_sp[net][freq].get(metric, 0.0)
                delta = abs(mv - bv)
                pf = "FAIL" if delta > thr else "WARN" if delta > thr * 0.8 else "PASS"
                results.append(
                    RegressionResult(
                        net_name=net,
                        metric=metric,
                        frequency_ghz=freq,
                        baseline_value=round(bv, 4),
                        modified_value=round(mv, 4),
                        delta=round(delta, 4),
                        threshold=thr,
                        pass_fail=pf,
                        impact=f"Δ={delta:+.3f} (thr={thr})",
                    )
                )
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────


def _hdr(cell, text: str, bg: str = _P["header_bg"]):
    cell.value = text
    cell.fill = _fill(bg)
    cell.font = Font(bold=True, color=_P["header_fg"], size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _dat(cell, value: Any, bg: Optional[str] = None, bold: bool = False):
    cell.value = value
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    cell.border = _thin_border()
    if bg:
        cell.fill = _fill(bg)
    if bold:
        cell.font = Font(bold=True)


def _row_bg(change_type: str, severity: str) -> Optional[str]:
    if change_type == ChangeType.ADDED.value:
        return _P["added_bg"]
    if change_type == ChangeType.REMOVED.value:
        return _P["removed_bg"]
    if change_type == ChangeType.MODIFIED.value:
        sev_col = {Severity.CRITICAL.value: _P["critical_bg"], Severity.MAJOR.value: _P["major_bg"]}
        return sev_col.get(severity, _P["modified_bg"])
    return None


def _autofit(ws, mn: int = 8, mx: int = 60):
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        best = mn
        for c in col:
            if c.value:
                best = min(mx, max(best, len(str(c.value)) + 2))
        ws.column_dimensions[ltr].width = best


# ── Sheet builders ─────────────────────────────────────────────────────────


def _sheet_cover(wb: Workbook, report: DiffReport):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "EDB Design Diff & Regression Report  ·  v3.0"
    c.font = Font(bold=True, size=16, color="FFFFFF")
    c.fill = _fill(_P["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 36

    meta = [
        ("Generated", report.generated_at),
        ("Run ID", report.run_id or "—"),
        ("Baseline EDB", report.baseline_path),
        ("Modified EDB", report.modified_path),
        ("Total Changes", len(report.diffs)),
        ("EM Sim Required", report.em_required_count),
        ("Elapsed (s)", report.elapsed_sec),
        ("Errors", len(report.errors)),
    ]
    for i, (lbl, val) in enumerate(meta, start=4):
        ws[f"B{i}"].value = lbl
        ws[f"B{i}"].font = Font(bold=True)
        ws[f"C{i}"].value = str(val)
        if lbl == "EM Sim Required" and report.em_required_count > 0:
            ws[f"C{i}"].fill = _fill(_P["em_req_bg"])
            ws[f"C{i}"].font = Font(bold=True, color="FFFFFF")

    # Git info block
    if report.git_info and report.git_info.current_commit:
        gi = report.git_info
        ws["E4"].value = "Git Branch"
        ws["E4"].font = Font(bold=True)
        ws["F4"].value = gi.current_branch
        ws["E5"].value = "Git Commit"
        ws["E5"].font = Font(bold=True)
        ws["F5"].value = gi.current_commit
        ws["E6"].value = "Commit Msg"
        ws["E6"].font = Font(bold=True)
        ws["F6"].value = gi.commit_message[:80]
        ws["E7"].value = "Dirty Tree"
        ws["E7"].font = Font(bold=True)
        ws["F7"].value = "Yes" if gi.is_dirty else "No"
        if gi.is_dirty:
            ws["F7"].fill = _fill(_P["warn_bg"])
        for r in range(4, 8):
            ws[f"E{r}"].fill = _fill(_P["git_bg"])
            ws[f"E{r}"].font = Font(bold=True, color="FFFFFF")

    ws["B13"].value = "Severity"
    ws["B13"].font = Font(bold=True)
    sev_bg = {
        Severity.CRITICAL.value: _P["critical_bg"],
        Severity.MAJOR.value: _P["major_bg"],
        Severity.MINOR.value: _P["pass_bg"],
    }
    for i, (sv, cnt) in enumerate(report.severity_counts.items(), start=14):
        ws[f"B{i}"].value = sv
        c2 = ws[f"C{i}"]
        c2.value = cnt
        c2.fill = _fill(sev_bg.get(sv, "FFFFFF"))
        c2.font = Font(bold=True, color="FFFFFF")

    ws["E13"].value = "Regression"
    ws["E13"].font = Font(bold=True)
    pf_bg = {"PASS": _P["pass_bg"], "WARN": _P["warn_bg"], "FAIL": _P["fail_bg"]}
    for i, (pf, cnt) in enumerate(report.regression_summary.items(), start=14):
        ws[f"E{i}"].value = pf
        c2 = ws[f"F{i}"]
        c2.value = cnt
        c2.fill = _fill(pf_bg.get(pf, "FFFFFF"))
        c2.font = Font(bold=True, color="FFFFFF")

    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 32


def _sheet_summary(wb: Workbook, report: DiffReport):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    hdrs = ["Category – Change Type", "Count"]
    ws.append(hdrs)
    for cell in ws[1]:
        _hdr(cell, cell.value)

    for key, cnt in report.summary.items():
        ct = key.split("–")[-1].strip() if "–" in key else ""
        ws.append([key, cnt])
        bg = _row_bg(ct, Severity.MINOR.value)
        for c in ws[ws.max_row]:
            if bg:
                c.fill = _fill(bg)

    # Bar chart
    n = len(report.summary)
    if n > 0:
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Change Count by Category"
        chart.y_axis.title = "Count"
        data = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 22
        chart.height = 14
        ws.add_chart(chart, "D2")

    _autofit(ws)


def _sheet_all_changes(wb: Workbook, report: DiffReport):
    ws = wb.create_sheet("All Changes")
    hdrs = [
        "#",
        "Category",
        "Change Type",
        "Severity",
        "Item Name",
        "Baseline Value",
        "Modified Value",
        "Field",
        "Impact Summary",
        "EM Sim Required",
        "EM Reasons",
    ]
    ws.append(hdrs)
    for i, cell in enumerate(ws[1], 1):
        _hdr(cell, hdrs[i - 1])
    ws.row_dimensions[1].height = 30

    for ri, entry in enumerate(report.diffs, start=2):
        em_bg = _P["em_req_bg"] if entry.em_sim_required else None
        bg = em_bg or _row_bg(entry.change_type, entry.severity)
        em_reasons_str = "; ".join(entry.em_sim_reasons) if entry.em_sim_reasons else ""
        row = [
            ri - 1,
            entry.category,
            entry.change_type,
            entry.severity,
            entry.item_name,
            str(entry.baseline_value),
            str(entry.modified_value),
            entry.detail,
            entry.impact_summary,
            "⚡ YES" if entry.em_sim_required else "No",
            em_reasons_str,
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            _dat(cell, val, bg=bg if bg else (_P["alt_row"] if ri % 2 == 0 else None))
            if ci == 10 and entry.em_sim_required:
                cell.fill = _fill(_P["em_req_bg"])
                cell.font = Font(bold=True, color="FFFFFF")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autofit(ws)


def _sheet_em_critical(wb: Workbook, report: DiffReport):
    """Dedicated sheet for all EM-simulation-required changes."""
    em_entries = [d for d in report.diffs if d.em_sim_required]
    if not em_entries:
        return
    ws = wb.create_sheet("⚡ EM Sim Required")
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:I1")
    banner = ws["A1"]
    banner.value = (
        f"⚡  {len(em_entries)} change(s) require EM re-simulation  "
        f"– validate with S-parameter / field solver before sign-off"
    )
    banner.font = Font(bold=True, size=12, color="FFFFFF")
    banner.fill = _fill(_P["em_req_bg"])
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    hdrs = ["#", "Category", "Item Name", "Field", "Baseline → Modified", "Severity", "EM Reasons", "Impact Summary"]
    ws.append(hdrs)
    for i, cell in enumerate(ws[2], 1):
        _hdr(cell, hdrs[i - 1], bg=_P["em_req_bg"])

    for ri, entry in enumerate(em_entries, start=3):
        change_str = f"{entry.baseline_value} → {entry.modified_value}"
        row = [
            ri - 2,
            entry.category,
            entry.item_name,
            entry.detail,
            change_str,
            entry.severity,
            "; ".join(entry.em_sim_reasons),
            entry.impact_summary,
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            _dat(cell, val, bg=_P["alt_row"] if ri % 2 == 0 else None)
            if ci == 6:
                sev_bg = {Severity.CRITICAL.value: _P["critical_bg"], Severity.MAJOR.value: _P["major_bg"]}
                bg = sev_bg.get(entry.severity)
                if bg:
                    cell.fill = _fill(bg)
                    cell.font = Font(bold=True, color="FFFFFF")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"
    _autofit(ws)


def _sheet_per_category(wb: Workbook, report: DiffReport):
    for cat in sorted({d.category for d in report.diffs}):
        ws = wb.create_sheet(cat[:31])
        hdrs = [
            "Change Type",
            "Severity",
            "Item Name",
            "Baseline Value",
            "Modified Value",
            "Field",
            "Impact Summary",
            "EM Sim Required",
        ]
        ws.append(hdrs)
        for i, cell in enumerate(ws[1], 1):
            _hdr(cell, hdrs[i - 1])

        for entry in [d for d in report.diffs if d.category == cat]:
            bg = _P["em_req_bg"] if entry.em_sim_required else _row_bg(entry.change_type, entry.severity)
            ri = ws.max_row + 1
            for ci, val in enumerate(
                [
                    entry.change_type,
                    entry.severity,
                    entry.item_name,
                    str(entry.baseline_value),
                    str(entry.modified_value),
                    entry.detail,
                    entry.impact_summary,
                    "⚡ YES" if entry.em_sim_required else "No",
                ],
                1,
            ):
                cell = ws.cell(ri, ci)
                _dat(cell, val, bg=bg)
                if ci == 8 and entry.em_sim_required:
                    cell.fill = _fill(_P["em_req_bg"])
                    cell.font = Font(bold=True, color="FFFFFF")

        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        _autofit(ws)


def _sheet_regression(wb: Workbook, report: DiffReport):
    if not report.regression:
        return
    ws = wb.create_sheet("Regression")
    hdrs = ["Net", "Metric", "Freq (GHz)", "Baseline", "Modified", "Δ Delta", "Threshold", "Pass/Fail", "Impact"]
    ws.append(hdrs)
    for i, cell in enumerate(ws[1], 1):
        _hdr(cell, hdrs[i - 1])
    ws.row_dimensions[1].height = 28

    sorted_reg = sorted(
        report.regression, key=lambda r: (r.pass_fail != "FAIL", r.pass_fail != "WARN", r.net_name, r.frequency_ghz)
    )
    pf_bg = {"FAIL": _P["fail_bg"], "WARN": _P["warn_bg"], "PASS": _P["pass_bg"]}

    for ri, res in enumerate(sorted_reg, start=2):
        row = [
            res.net_name,
            res.metric,
            res.frequency_ghz,
            res.baseline_value,
            res.modified_value,
            res.delta,
            res.threshold,
            res.pass_fail,
            res.impact,
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            bg = pf_bg.get(res.pass_fail) if ci == 8 else (_P["alt_row"] if ri % 2 == 0 else None)
            _dat(cell, val, bg=bg, bold=(ci == 8))
            if ci == 8:
                cell.font = Font(bold=True, color="FFFFFF")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autofit(ws)


def _sheet_git(wb: Workbook, report: DiffReport):
    """Git information and recent commit log."""
    gi = report.git_info
    if gi is None or not gi.current_commit:
        return
    ws = wb.create_sheet("Git")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Git Repository Information"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _fill(_P["git_bg"])
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    rows = [
        ("Repository", gi.repo_path),
        ("Branch", gi.current_branch),
        ("Commit", gi.current_commit),
        ("Message", gi.commit_message),
        ("Dirty Tree", "Yes – uncommitted changes present" if gi.is_dirty else "No"),
    ]
    if gi.new_commit_hash:
        rows.append(("Auto-Committed", gi.new_commit_hash))

    for i, (lbl, val) in enumerate(rows, start=3):
        ws.cell(i, 1).value = lbl
        ws.cell(i, 1).font = Font(bold=True)
        ws.cell(i, 2).value = val
        if lbl == "Dirty Tree" and gi.is_dirty:
            ws.cell(i, 2).fill = _fill(_P["warn_bg"])

    if gi.changed_files:
        r = ws.max_row + 2
        ws.cell(r, 1).value = "Changed / Untracked Files"
        ws.cell(r, 1).font = Font(bold=True)
        for f in gi.changed_files[:50]:
            r += 1
            ws.cell(r, 1).value = f

    _autofit(ws)


def _sheet_baseline_validations(wb: Workbook, report: DiffReport):
    if not report.baseline_validations:
        return
    ws = wb.create_sheet("Baseline Validation")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = "Baseline Reference Validation Suite"
    c.font = Font(bold=True, size=13, color="FFFFFF")
    c.fill = _fill(_P["title_bg"])
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 26

    hdrs = [
        "Baseline Name",
        "Tag",
        "Stored At",
        "Total Changes",
        "Critical",
        "Major",
        "EM Required",
        "Reg Fails",
        "Result",
    ]
    ws.append(hdrs)
    for i, cell in enumerate(ws[2], 1):
        _hdr(cell, hdrs[i - 1])

    for bv in report.baseline_validations:
        ri = ws.max_row + 1
        result_str = "✔ PASS" if bv.passed else "✘ FAIL"
        row = [
            bv.baseline_name,
            bv.baseline_tag,
            bv.stored_at,
            bv.total_changes,
            bv.critical_count,
            bv.major_count,
            bv.em_required_count,
            bv.regression_fails,
            result_str,
        ]
        for ci, val in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            bg = (
                (_P["baseline_ok"] if bv.passed else _P["baseline_fail"])
                if ci == 9
                else (_P["alt_row"] if ri % 2 == 0 else None)
            )
            _dat(cell, val, bg=bg)
            if ci == 9:
                cell.font = Font(bold=True, color="FFFFFF")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A3"
    _autofit(ws)


def _sheet_errors(wb: Workbook, report: DiffReport):
    if not report.errors:
        return
    ws = wb.create_sheet("Errors")
    hdrs = ["Phase", "Message", "Traceback"]
    ws.append(hdrs)
    for i, cell in enumerate(ws[1], 1):
        _hdr(cell, cell.value, bg=_P["critical_bg"])
    for e in report.errors:
        ws.append([e.phase, e.message, e.tb[:300]])
    _autofit(ws, mx=80)


def write_excel_report(report: DiffReport, output_path: str):
    wb = Workbook()
    _sheet_cover(wb, report)
    _sheet_summary(wb, report)
    _sheet_all_changes(wb, report)
    _sheet_em_critical(wb, report)  # v3: EM criticality sheet
    _sheet_per_category(wb, report)
    _sheet_regression(wb, report)
    _sheet_git(wb, report)  # v3: Git info sheet
    _sheet_baseline_validations(wb, report)  # v3: baseline validation sheet
    _sheet_errors(wb, report)
    wb.save(output_path)
    console.print(f"\n[bold green]✔ Excel report → {output_path}[/bold green]")


# ─────────────────────────────────────────────────────────────────────────────
# JSON export
# ─────────────────────────────────────────────────────────────────────────────


def write_json_report(report: DiffReport, output_path: str):
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(report.to_dict(), f, indent=2, default=str)
    console.print(f"[bold green]✔ JSON report → {output_path}[/bold green]")


# ─────────────────────────────────────────────────────────────────────────────
# Console summary
# ─────────────────────────────────────────────────────────────────────────────


def print_console_summary(report: DiffReport):
    console.rule("📋  Diff Summary")

    if HAS_RICH:
        tbl = Table(show_header=True, header_style="bold white on blue", border_style="dim", expand=False)
        tbl.add_column("Category – Change", width=38)
        tbl.add_column("Count", justify="right", width=8)
        for key, cnt in report.summary.items():
            clr = "green" if "Added" in key else "red" if "Removed" in key else "yellow"
            tbl.add_row(key, f"[{clr}]{cnt}[/{clr}]")
        console.print(tbl)

        sev_tbl = Table(show_header=True, header_style="bold", border_style="dim", expand=False, title="Severity")
        sev_tbl.add_column("Level", width=12)
        sev_tbl.add_column("Count", justify="right", width=8)
        sev_clr = {Severity.CRITICAL.value: "red", Severity.MAJOR.value: "yellow", Severity.MINOR.value: "green"}
        for sv, cnt in report.severity_counts.items():
            c = sev_clr.get(sv, "white")
            sev_tbl.add_row(f"[{c}]{sv}[/{c}]", str(cnt))
        console.print(sev_tbl)
    else:
        for key, cnt in report.summary.items():
            print(f"  {key}: {cnt}")

    # v3: EM criticality summary
    if report.em_required_count > 0:
        console.rule("⚡  EM Simulation Required")
        if HAS_RICH:
            em_tbl = Table(show_header=True, header_style="bold white on purple", border_style="dim", expand=False)
            em_tbl.add_column("Item", width=28)
            em_tbl.add_column("Field / Type", width=20)
            em_tbl.add_column("EM Reasons", width=52)
            for e in report.em_critical_entries[:20]:
                em_tbl.add_row(
                    f"[bold]{e.item_name}[/bold]",
                    e.detail or e.change_type,
                    "[dim]" + "; ".join(e.em_sim_reasons) + "[/dim]",
                )
            console.print(em_tbl)
            if len(report.em_critical_entries) > 20:
                console.print(f"  … and {len(report.em_critical_entries) - 20} more (see Excel)")
        else:
            print(f"  EM re-simulation required for {report.em_required_count} change(s).")

    # v3: Git info
    if report.git_info and report.git_info.current_commit:
        gi = report.git_info
        console.rule("🔀  Git")
        console.print(
            f"  Branch: [bold]{gi.current_branch}[/bold]  "
            f"Commit: [bold]{gi.current_commit}[/bold]  "
            f"{'[red]Dirty[/red]' if gi.is_dirty else '[green]Clean[/green]'}"
        )
        if gi.new_commit_hash:
            console.print(f"  Auto-committed → [bold]{gi.new_commit_hash}[/bold]")

    # v3: DB run ID
    if report.run_id:
        console.print(f"  🗄  Saved to DB – run_id={report.run_id}")

    if report.regression:
        console.rule("📡  Regression")
        rs = report.regression_summary
        console.print(
            f"  PASS [green]{rs['PASS']}[/green]  WARN [yellow]{rs['WARN']}[/yellow]  FAIL [red]{rs['FAIL']}[/red]"
        )
        fails = [r for r in report.regression if r.pass_fail == "FAIL"]
        if fails and HAS_RICH:
            ft = Table(
                show_header=True, header_style="bold red", border_style="dim", expand=False, title="FAILed checks"
            )
            for col in ("Net", "Metric", "Freq GHz", "Baseline", "Modified", "Δ", "Thr"):
                ft.add_column(col)
            for r in fails[:25]:
                ft.add_row(
                    r.net_name,
                    r.metric,
                    str(r.frequency_ghz),
                    str(r.baseline_value),
                    str(r.modified_value),
                    f"[red]{r.delta}[/red]",
                    str(r.threshold),
                )
            console.print(ft)

    # v3: Baseline validation summary
    if report.baseline_validations:
        console.rule("🎯  Baseline Validation Suite")
        if HAS_RICH:
            bv_tbl = Table(show_header=True, header_style="bold", border_style="dim", expand=False)
            bv_tbl.add_column("Baseline", width=24)
            bv_tbl.add_column("Changes", justify="right", width=10)
            bv_tbl.add_column("Critical", justify="right", width=10)
            bv_tbl.add_column("EM Req.", justify="right", width=10)
            bv_tbl.add_column("Result", width=10)
            for bv in report.baseline_validations:
                res_clr = "green" if bv.passed else "red"
                res_str = "✔ PASS" if bv.passed else "✘ FAIL"
                bv_tbl.add_row(
                    bv.baseline_name,
                    str(bv.total_changes),
                    str(bv.critical_count),
                    str(bv.em_required_count),
                    f"[{res_clr}]{res_str}[/{res_clr}]",
                )
            console.print(bv_tbl)

    if report.errors:
        console.print(f"\n[bold red]⚠  {len(report.errors)} extraction error(s) – check 'Errors' sheet.[/bold red]")
    console.print(f"\n  ⏱  Completed in {report.elapsed_sec}s\n")


def print_run_history(db: DiffDatabase, limit: int = 20):
    """Print a formatted table of past runs from the database."""
    rows = db.get_run_history(limit=limit)
    if not rows:
        console.print("[yellow]No runs found in database.[/yellow]")
        return
    console.rule(f"🗄  Run History ({len(rows)} entries)")
    if HAS_RICH:
        tbl = Table(show_header=True, header_style="bold white on dark_blue", border_style="dim", expand=False)
        for col in ("ID", "Date", "Baseline", "Modified", "Diffs", "Critical", "EM Req", "Fails", "Branch"):
            tbl.add_column(col)
        for r in rows:
            fail_str = str(r["reg_fail"])
            fail_clr = "red" if r["reg_fail"] > 0 else "green"
            crit_clr = "red" if r["critical_count"] > 0 else "green"
            em_clr = "purple" if r["em_required"] > 0 else "green"
            tbl.add_row(
                str(r["id"]),
                r["generated_at"][:16],
                Path(r["baseline_path"]).name if r["baseline_path"] else "—",
                Path(r["modified_path"]).name if r["modified_path"] else "—",
                str(r["total_diffs"]),
                f"[{crit_clr}]{r['critical_count']}[/{crit_clr}]",
                f"[{em_clr}]{r['em_required']}[/{em_clr}]",
                f"[{fail_clr}]{fail_str}[/{fail_clr}]",
                r["git_branch"] or "—",
            )
        console.print(tbl)
    else:
        for r in rows:
            print(
                f"  [{r['id']}] {r['generated_at'][:16]}  "
                f"diffs={r['total_diffs']}  critical={r['critical_count']}  "
                f"em={r['em_required']}  branch={r['git_branch'] or '—'}"
            )


def print_baseline_list(db: DiffDatabase):
    """Print all stored baseline references."""
    refs = db.list_baseline_references()
    if not refs:
        console.print("[yellow]No baseline references stored in database.[/yellow]")
        return
    console.rule("🎯  Stored Baseline References")
    if HAS_RICH:
        tbl = Table(show_header=True, header_style="bold white on green", border_style="dim", expand=False)
        for col in ("ID", "Name", "Tag", "Stored At", "Git Commit", "Branch", "Notes"):
            tbl.add_column(col)
        for r in refs:
            tbl.add_row(
                str(r["id"]),
                r["name"],
                r.get("tag", ""),
                r["stored_at"][:16],
                r.get("git_commit", "")[:12],
                r.get("git_branch", ""),
                r.get("notes", "")[:40],
            )
        console.print(tbl)
    else:
        for r in refs:
            print(f"  [{r['id']}] {r['name']}  {r['stored_at'][:16]}  commit={r.get('git_commit', '')[:12]}")


# ─────────────────────────────────────────────────────────────────────────────
# Demo data (no EDB files needed)
# ─────────────────────────────────────────────────────────────────────────────


def _demo_report() -> DiffReport:
    report = DiffReport(
        baseline_path="baseline_demo.aedb",
        modified_path="modified_demo.aedb",
        elapsed_sec=0.61,
        git_info=GitInfo(
            repo_path="/demo/project",
            current_branch="feature/stackup-update",
            current_commit="a1b2c3d4e5f6",
            commit_message="WIP: update dielectric stack for 28GHz",
            is_dirty=True,
            changed_files=["design/board.aedb", "stackup/layer_config.yaml"],
        ),
    )
    report.diffs = [
        # Nets
        DiffEntry(
            "Net",
            "Added",
            "VDD_1V8",
            "—",
            "component_count=6",
            severity="Major",
            impact_summary="New power net VDD_1V8 introduced.",
        ),
        DiffEntry(
            "Net",
            "Removed",
            "VDD_1V5",
            "component_count=6",
            "—",
            severity="Major",
            impact_summary="Power net VDD_1V5 removed.",
        ),
        DiffEntry(
            "Net",
            "Modified",
            "GND",
            120,
            122,
            detail="component_count",
            severity="Minor",
            impact_summary="'component_count' changed by +2 (+1.7%).",
        ),
        # Components
        DiffEntry(
            "Component",
            "Added",
            "U12",
            "—",
            "part_name=MCU_v2",
            severity="Major",
            impact_summary="New component U12 added.",
        ),
        DiffEntry(
            "Component",
            "Removed",
            "U11",
            "part_name=MCU_v1",
            "—",
            severity="Major",
            impact_summary="Component U11 removed.",
        ),
        DiffEntry(
            "Component",
            "Modified",
            "C3",
            5.0,
            5.25,
            detail="center_x_mm",
            severity="Minor",
            impact_summary="'center_x_mm' changed by +0.25 (+5.0%).",
        ),
        DiffEntry(
            "Component",
            "Modified",
            "R7",
            "10k",
            "22k",
            detail="value",
            severity="Major",
            impact_summary="'value' changed: '10k' → '22k'.",
        ),
        DiffEntry(
            "Component",
            "Modified",
            "U3",
            "BGA_144",
            "BGA_196",
            detail="part_name",
            severity="Major",
            impact_summary="'part_name' changed: 'BGA_144' → 'BGA_196'.",
        ),
        # Layers
        DiffEntry(
            "Layer",
            "Modified",
            "Signal_L2",
            35,
            30,
            detail="thickness_um",
            severity="Major",
            impact_summary="'thickness_um' changed by -5 (-14.3%).",
        ),
        DiffEntry(
            "Layer",
            "Modified",
            "Diel_D1",
            100,
            90,
            detail="thickness_um",
            severity="Major",
            impact_summary="'thickness_um' changed by -10 (-10.0%).",
        ),
        DiffEntry(
            "Layer",
            "Modified",
            "Diel_D1",
            4.3,
            3.9,
            detail="Er",
            severity="Major",
            impact_summary="'Er' changed by -0.4 (-9.3%).",
        ),
        DiffEntry(
            "Layer",
            "Modified",
            "Diel_D1",
            0.02,
            0.025,
            detail="loss_tangent",
            severity="Major",
            impact_summary="'loss_tangent' changed by +0.005 (+25.0%).",
        ),
        # Padstacks
        DiffEntry(
            "Padstack",
            "Modified",
            "Via_0.2mm",
            0.20,
            0.15,
            detail="drill_mm",
            severity="Major",
            impact_summary="'drill_mm' changed by -0.05 (-25.0%).",
        ),
        DiffEntry(
            "Padstack",
            "Added",
            "Via_0.1mm",
            "—",
            "drill_mm=0.10",
            severity="Major",
            impact_summary="New padstack Via_0.1mm introduced.",
        ),
        # Traces
        DiffEntry(
            "Trace",
            "Modified",
            "DIFF_PAIR_A",
            0.12,
            0.10,
            detail="min_width_mm",
            severity="Major",
            impact_summary="'min_width_mm' changed by -0.02 (-16.7%).",
        ),
        DiffEntry(
            "Trace",
            "Modified",
            "CLK_NET",
            52.3,
            48.1,
            detail="total_length_mm",
            severity="Minor",
            impact_summary="'total_length_mm' changed by -4.2 (-8.0%).",
        ),
        # Vias
        DiffEntry(
            "Via",
            "Modified",
            "GND",
            48,
            52,
            detail="via_count",
            severity="Minor",
            impact_summary="'via_count' changed by +4 (+8.3%).",
        ),
        DiffEntry(
            "Via",
            "Modified",
            "VCC",
            12,
            10,
            detail="via_count",
            severity="Minor",
            impact_summary="'via_count' changed by -2 (-16.7%).",
        ),
    ]

    # Apply EM criticality to demo diffs
    cfg = DiffConfig()
    report.em_critical_entries = apply_em_criticality(report.diffs, cfg)

    report.regression = [
        RegressionResult("VDD_1V8", "IL_dB", 1.0, -0.5, -0.6, 0.10, 2.0, "PASS"),
        RegressionResult("VDD_1V8", "IL_dB", 5.0, -1.2, -1.5, 0.30, 2.0, "PASS"),
        RegressionResult("VDD_1V8", "IL_dB", 10.0, -2.1, -4.5, 2.40, 2.0, "FAIL", impact="Δ=+2.400 (thr=2.0)"),
        RegressionResult("GND", "RL_dB", 1.0, -20.0, -18.5, 1.50, 3.0, "PASS"),
        RegressionResult("GND", "RL_dB", 2.0, -18.0, -15.6, 2.40, 3.0, "WARN", impact="Δ=+2.400 (80% warn)"),
        RegressionResult("GND", "RL_dB", 5.0, -15.0, -11.0, 4.00, 3.0, "FAIL", impact="Δ=+4.000 (thr=3.0)"),
        RegressionResult("DIFF_PAIR_A", "Impedance_Ohm", 1.0, 100.0, 103.5, 3.5, 5.0, "PASS"),
        RegressionResult(
            "DIFF_PAIR_A", "Impedance_Ohm", 5.0, 100.0, 108.0, 8.0, 5.0, "FAIL", impact="Δ=+8.000 (thr=5.0)"
        ),
        RegressionResult("CLK_NET", "IL_dB", 3.0, -0.8, -1.1, 0.30, 2.0, "PASS"),
    ]

    # Demo baseline validations
    report.baseline_validations = [
        BaselineValidationResult(
            baseline_name="golden_v1.0",
            baseline_tag="release",
            stored_at="2024-11-01T10:00:00",
            total_changes=18,
            critical_count=3,
            major_count=9,
            minor_count=6,
            em_required_count=4,
            regression_fails=3,
            passed=False,
            summary="FAIL – 18 total change(s), 3 critical, 4 EM-required",
        ),
        BaselineValidationResult(
            baseline_name="golden_v0.9",
            baseline_tag="pre-release",
            stored_at="2024-10-01T09:00:00",
            total_changes=5,
            critical_count=0,
            major_count=3,
            minor_count=2,
            em_required_count=0,
            regression_fails=0,
            passed=True,
            summary="PASS – 5 total change(s), 0 critical, 0 EM-required",
        ),
    ]
    return report


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────


def _parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="PyEDB Local Design Diff & Regression Tool v3.0",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    # Core diff
    p.add_argument("--baseline", help="Path to baseline .aedb")
    p.add_argument("--modified", help="Path to modified .aedb")
    p.add_argument(
        "--report", default="design_diff_report.xlsx", help="Excel output (default: design_diff_report.xlsx)"
    )
    p.add_argument("--json", default="", help="Optional JSON output path")
    p.add_argument("--config", default="", help="Optional YAML config file")
    p.add_argument("--simulate", action="store_true", help="Run S-param regression (requires AEDT licence)")
    p.add_argument("--freq-ghz", type=float, default=10.0, help="Max simulation frequency in GHz (default: 10)")
    p.add_argument("--edb-version", default="2024.1", help="AEDT version string (default: 2024.1)")
    p.add_argument("--demo", action="store_true", help="Synthetic demo data (no EDB files required)")
    p.add_argument("--verbose", action="store_true", help="Enable DEBUG logging")
    # v3: Git
    p.add_argument("--git-repo", default="", help="Path to Git repository root (enables Git integration)")
    p.add_argument("--git-commit", action="store_true", help="Auto-commit dirty working tree before diff")
    # v3: Database
    p.add_argument("--db", default="", help="Path to SQLite DB file (enables persistent storage)")
    p.add_argument("--history", action="store_true", help="Print run history from DB and exit")
    p.add_argument("--list-baselines", action="store_true", help="List stored baseline references and exit")
    # v3: Baseline reference management
    p.add_argument(
        "--save-baseline", default="", metavar="NAME", help="Save --modified design as a named baseline reference in DB"
    )
    p.add_argument("--baseline-tag", default="", help="Tag for --save-baseline (e.g. 'v1.2')")
    p.add_argument("--baseline-notes", default="", help="Notes for --save-baseline")
    p.add_argument(
        "--validate-baselines", action="store_true", help="Validate --modified against all stored baselines in DB"
    )
    p.add_argument(
        "--baseline-names", nargs="*", default=None, help="Filter --validate-baselines to specific named baselines"
    )
    # Utility
    p.add_argument(
        "--export-config",
        default="",
        metavar="PATH",
        help="Write current effective config to a YAML file and exit (useful for bootstrapping a config)",
    )
    return p


def main():
    args = _parser().parse_args()
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    cfg = DiffConfig()
    if args.config and Path(args.config).exists():
        cfg = DiffConfig.from_yaml(args.config)
    cfg.edb_version = args.edb_version
    cfg.sim_stop_freq = f"{args.freq_ghz}GHz"
    if args.db:
        cfg.db_path = args.db

    # ── Export config and exit ──────────────────────────────────────────────
    if args.export_config:
        if not HAS_YAML:
            sys.exit("--export-config requires pyyaml:  pip install pyyaml")
        cfg.to_yaml(args.export_config)
        console.print(f"[bold green]✔ Config exported → {args.export_config}[/bold green]")
        return

    # ── Open DB if requested ────────────────────────────────────────────────
    db: Optional[DiffDatabase] = None
    if args.db:
        db = DiffDatabase(args.db)

    # ── History / list mode ─────────────────────────────────────────────────
    if args.history:
        if db is None:
            sys.exit("--history requires --db <path>")
        print_run_history(db)
        db.close()
        return

    if args.list_baselines:
        if db is None:
            sys.exit("--list-baselines requires --db <path>")
        print_baseline_list(db)
        db.close()
        return

    # ── Git integration ─────────────────────────────────────────────────────
    git_info: Optional[GitInfo] = None
    if args.git_repo:
        cfg.git_auto_commit = args.git_commit
        git_info = _capture_git_info(
            args.git_repo,
            auto_commit=cfg.git_auto_commit,
            commit_msg=cfg.git_commit_message,
        )

    # ── Save baseline reference ─────────────────────────────────────────────
    if args.save_baseline:
        if not args.modified:
            sys.exit("--save-baseline requires --modified <path>")
        if db is None:
            sys.exit("--save-baseline requires --db <path>")
        save_baseline_reference(
            db=db,
            name=args.save_baseline,
            design_path=args.modified,
            cfg=cfg,
            git_info=git_info,
            tag=args.baseline_tag,
            notes=args.baseline_notes,
        )
        db.close()
        return

    # ── Validate against baselines ──────────────────────────────────────────
    if args.validate_baselines:
        if not args.modified:
            sys.exit("--validate-baselines requires --modified <path>")
        if db is None:
            sys.exit("--validate-baselines requires --db <path>")
        validations, sub_reports = validate_against_baselines(
            modified_path=args.modified,
            db=db,
            cfg=cfg,
            baseline_names=args.baseline_names,
        )
        # Build a combined report for output
        combined = DiffReport(
            baseline_path="(multiple baselines)",
            modified_path=args.modified,
            git_info=git_info,
        )
        combined.baseline_validations = validations
        for sr in sub_reports:
            combined.diffs.extend(sr.diffs)
        combined.em_critical_entries = [d for d in combined.diffs if d.em_sim_required]
        if db:
            combined.run_id = db.save_run(combined)
        print_console_summary(combined)
        write_excel_report(combined, args.report)
        if args.json:
            write_json_report(combined, args.json)
        db.close()
        return

    # ── Normal diff mode ─────────────────────────────────────────────────────
    if args.demo:
        if HAS_RICH:
            console.print(
                Panel("[bold yellow]DEMO MODE[/bold yellow] – synthetic data, no real EDB files needed.", expand=False)
            )
        else:
            print("=== DEMO MODE ===")
        report = _demo_report()
    else:
        if not args.baseline or not args.modified:
            _parser().error("--baseline and --modified are required unless --demo is set")
        report = run_diff(args.baseline, args.modified, cfg=cfg, db=db, git_info=git_info)
        if args.simulate:
            console.rule("Running Simulation Regression")
            report.regression = run_simulation_regression(args.baseline, args.modified, cfg=cfg)

    print_console_summary(report)
    write_excel_report(report, args.report)
    if args.json:
        write_json_report(report, args.json)

    if db is not None:
        db.close()


if __name__ == "__main__":
    main()
